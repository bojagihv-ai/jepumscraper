import logging
import os
import shutil
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageOps

from scrapers.base_scraper import ProductResult

logger = logging.getLogger(__name__)


def _safe_preview_path(output_path: str, product_id: str, index: int) -> str:
    preview_dir = os.path.join(os.path.dirname(output_path), "_excel_previews")
    os.makedirs(preview_dir, exist_ok=True)
    safe_id = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(product_id))[:80]
    return os.path.join(preview_dir, f"{safe_id}_{index:02d}.jpg")


def _safe_asset_path(output_path: str, product_id: str, source_path: str, index: int) -> str:
    asset_dir = os.path.join(
        os.path.dirname(output_path),
        "_excel_assets",
        os.path.splitext(os.path.basename(output_path))[0],
    )
    os.makedirs(asset_dir, exist_ok=True)
    safe_id = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(product_id))[:80]
    ext = os.path.splitext(source_path)[1].lower() or ".jpg"
    return os.path.join(asset_dir, f"{safe_id}_{index:02d}{ext}")


def _copy_detail_asset(source_path: str, output_path: str, product_id: str, index: int) -> str:
    asset_path = _safe_asset_path(output_path, product_id, source_path, index)
    if os.path.abspath(source_path) != os.path.abspath(asset_path):
        shutil.copy2(source_path, asset_path)
    return asset_path


def _make_preview_image(source_path: str, preview_path: str, max_width: int = 190, max_height: int = 116) -> tuple[int, int]:
    """Create a compact embedded preview so tall screenshots do not break the sheet layout."""
    with PILImage.open(source_path) as img:
        img = ImageOps.exif_transpose(img).convert("RGB")
        width, height = img.size
        if height > width * 2:
            crop_h = min(height, max(900, int(width * 0.85)))
            img = img.crop((0, 0, width, crop_h))
        img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
        canvas = PILImage.new("RGB", (max_width, max_height), "white")
        x = (max_width - img.width) // 2
        y = (max_height - img.height) // 2
        canvas.paste(img, (x, y))
        canvas.save(preview_path, "JPEG", quality=84)
        return canvas.size


def _make_thumbnail_image(source_path: str, preview_path: str, size: int = 88) -> tuple[int, int]:
    """Write a real thumbnail file; some spreadsheet apps ignore display-only image sizing."""
    with PILImage.open(source_path) as img:
        img = ImageOps.exif_transpose(img).convert("RGB")
        img.thumbnail((size, size), PILImage.Resampling.LANCZOS)
        canvas = PILImage.new("RGB", (size, size), "white")
        x = (size - img.width) // 2
        y = (size - img.height) // 2
        canvas.paste(img, (x, y))
        canvas.save(preview_path, "JPEG", quality=86)
        return canvas.size


def _file_uri(path: str) -> str:
    try:
        return Path(path).resolve().as_uri()
    except Exception:
        return os.path.abspath(path).replace("\\", "/")


class ExcelExporter:
    def export(self, products: List[ProductResult], detail_data: Dict[str, Dict], output_path: str):
        logger.info("Exporting %s products to Excel: %s", len(products), output_path)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "경쟁사_제품_분석"

            max_detail_parts = max(
                [len((detail_data.get(prod.id, {}) or {}).get("screenshots") or []) for prod in products] or [0]
            )
            headers = ["매칭단계", "플랫폼", "업체명", "제품명", "가격", "URL", "썸네일", "상세 미리보기", "상세 원본 파일"]
            if max_detail_parts > 1:
                headers.extend([f"상세 원본 {idx}" for idx in range(2, max_detail_parts + 1)])
            last_col = len(headers)

            header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
            header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            widths = {"A": 13, "B": 13, "C": 18, "D": 44, "E": 12, "F": 58, "G": 18, "H": 28, "I": 54}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            for col_idx in range(10, last_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 28
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}1"

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            tier_colors = {1: "E53E3E", 2: "DD6B20", 3: "3182CE", 0: "718096"}
            tier_labels = {1: "1단계(동일)", 2: "2단계(유사)", 3: "3단계(색상/옵션)", 0: "미분류"}
            row_idx = 2

            for prod in products:
                ws.cell(row=row_idx, column=1, value=tier_labels.get(prod.match_tier, "?"))
                ws.cell(row=row_idx, column=2, value=prod.platform)
                ws.cell(row=row_idx, column=3, value=getattr(prod, "seller_name", "") or prod.platform)
                ws.cell(row=row_idx, column=4, value=prod.title)
                ws.cell(row=row_idx, column=5, value=prod.price)

                url_cell = ws.cell(row=row_idx, column=6, value=prod.product_url)
                url_cell.hyperlink = prod.product_url
                url_cell.font = Font(color="4299E1", underline="single")

                ws.row_dimensions[row_idx].height = 96
                ws.cell(row=row_idx, column=1).font = Font(
                    bold=True,
                    color=tier_colors.get(prod.match_tier, "718096"),
                )

                for col_idx in range(1, last_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        vertical="center",
                        wrap_text=col_idx in (1, 3, 4, 6) or col_idx >= 9,
                    )

                if prod.local_thumbnail_path and os.path.exists(prod.local_thumbnail_path):
                    try:
                        thumb_preview_path = _safe_preview_path(output_path, prod.id, 0)
                        thumb_w, thumb_h = _make_thumbnail_image(prod.local_thumbnail_path, thumb_preview_path)
                        img = ExcelImage(thumb_preview_path)
                        img.width = thumb_w
                        img.height = thumb_h
                        ws.add_image(img, f"G{row_idx}")
                    except Exception as exc:
                        logger.error("Cannot add thumbnail: %s", exc)

                product_data = detail_data.get(prod.id, {}) or {}
                screenshots = product_data.get("screenshots") or []
                valid_paths = [path for path in screenshots if path and os.path.exists(path)]
                if valid_paths:
                    first_asset_path = valid_paths[0]
                    try:
                        first_asset_path = _copy_detail_asset(valid_paths[0], output_path, prod.id, 1)
                        preview_path = _safe_preview_path(output_path, prod.id, 1)
                        preview_w, preview_h = _make_preview_image(first_asset_path, preview_path)
                        img_detail = ExcelImage(preview_path)
                        img_detail.width = preview_w
                        img_detail.height = preview_h
                        ws.add_image(img_detail, f"H{row_idx}")
                    except Exception as exc:
                        logger.error("Cannot add detail preview: %s", exc)

                    detail_cell = ws.cell(row=row_idx, column=9)
                    detail_cell.value = "상세 이미지 열기" if len(valid_paths) == 1 else f"상세 이미지 열기 외 {len(valid_paths) - 1}개"
                    detail_cell.hyperlink = _file_uri(first_asset_path)
                    detail_cell.font = Font(color="4299E1", underline="single", size=9)
                    detail_cell.alignment = Alignment(wrap_text=True, vertical="center")

                    for idx, part_path in enumerate(valid_paths[1:], start=2):
                        try:
                            asset_path = _copy_detail_asset(part_path, output_path, prod.id, idx)
                            img_col = 8 + idx
                            cell = ws.cell(row=row_idx, column=img_col)
                            cell.value = f"분할 {idx} 열기"
                            cell.hyperlink = _file_uri(asset_path)
                            cell.font = Font(color="4299E1", underline="single", size=9)
                            cell.alignment = Alignment(wrap_text=True, vertical="center")
                        except Exception as exc:
                            logger.error("Cannot add detail link: %s", exc)

                mhtml_path = product_data.get("mhtml_path", "")
                if mhtml_path:
                    mhtml_cell = ws.cell(row=row_idx, column=9)
                    if mhtml_cell.value:
                        mhtml_cell.value = f"{mhtml_cell.value}\nMHTML: {mhtml_path}"
                    else:
                        mhtml_cell.value = mhtml_path
                        mhtml_cell.font = Font(size=9, color="999999")
                    mhtml_cell.alignment = Alignment(wrap_text=True, vertical="center")

                row_idx += 1

            wb.save(output_path)
            logger.info("Excel file generated successfully.")
            return True

        except Exception as exc:
            logger.exception("Failed to export excel: %s", exc)
            return False
